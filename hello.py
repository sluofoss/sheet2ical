import openpyxl.worksheet
import openpyxl
import datetime
import csv
import icsconverter

def main(filepath = "./weekly timetable template.xlsx", export_path ="./weekly timetable template.ical"):
    start_monday, end_monday, events = get_event_from_timetable(filepath)
    events_to_ical(start_monday, end_monday, events, export_path)

def events_to_ical(start_monday, end_monday, events, ical_path):
    events_csv_prep = []
    for e, info in events.items():
        #print(type(info['start_time_in_day']))
        curr_monday = start_monday#datetime.datetime.strptime(start_monday, 'YYYY-MM-DD')
        flag = True
        while flag:
            for day in info['repeat_day']:
                events_csv_prep.append({
                    'Subject': info['val'],
                    'Start Date': datetime.datetime.strftime(curr_monday + datetime.timedelta(days = day),'%m/%d/%Y'),
                    'Start Time': str(info['start_time_in_day'])[:5], #datetime.datetime.strftime(info['start_time_in_day'], '%H:%M'),
                    'End Date': datetime.datetime.strftime(curr_monday + datetime.timedelta(days = day),'%m/%d/%Y'),
                    'End Time': str(info['end_time_in_day'])[:5], #datetime.datetime.strftime(info['end_time_in_day'], '%H:%M'),
                    'All Day Event': False,
                    'Description': ' ',
                    'Location': ' ',
                    'Private': ' ',
                })
            curr_monday += datetime.timedelta(days=7)
            # debug
            if curr_monday >= end_monday:
                break

    with open('tmp.csv', mode='w') as csv_file:
        fieldnames = events_csv_prep[0].keys() # ['emp_name', 'dept', 'birth_month']
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for r in events_csv_prep: 
            writer.writerow(r)

    icsconverter.main("tmp.csv")


        
def get_event_from_timetable(filepath):
    #print("Hello from sheet2ical!")
    
    wb_obj = openpyxl.load_workbook(filepath)
    
    day_start_index = None
    ws = wb_obj['Sheet1']
    for row in ws.iter_rows(min_row=3, min_col = 1, max_col=1, max_row=26):
        for cell in row:
            if cell.value <= datetime.time(0,0,0):
                # this is when the day actually starts
                #print(cell.value, cell.coordinate, cell.row, cell.column)
                day_start_index = {'row': cell.row, 'column': cell.column}
                break
    if day_start_index is None:
        raise Exception('time 00:00 missing from index')

    start_monday = ws.cell(row = 1, column = 2).value
    end_monday = ws.cell(row = 1, column = 4).value
    # start:end
    unique_val_cells = {}
    for row in ws.iter_rows(min_row=3, min_col = 3, max_col=9, max_row=26):
        for cell in row:
            if isinstance(cell,openpyxl.worksheet.merge.MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # return the left top cell
                        left_top_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        right_bottom_cell = ws.cell(row=merged_range.max_row, column=merged_range.max_col)
                        break
                #print('found cell', left_top_cell, right_bottom_cell)
                #print(left_top_cell.value, right_bottom_cell.value)
                #print(left_top_cell.coordinate, right_bottom_cell.coordinate)
                #print('is left already observed?', left_top_cell in unique_val_cells)
                unique_val_cells[left_top_cell] = right_bottom_cell
            else:
                #print('info')
                #print(cell, cell.value)
                #print('info end')
                unique_val_cells[cell] = None
                # print(cell.coordinate)
                #print(cell.__dir__())
    event_info = {}
    for cell in unique_val_cells.items():   
        item_start_day =  cell[0].column - 3
        item_start_time = ws.cell(column = 1, row = cell[0].row).value
        
        item_start_last_night = False

        
        if cell[0].row < day_start_index['row']:
            #item_start_day = item_start_day-1
            item_start_last_night = True
        
        if cell[1] is None:
            item_end_time = ws.cell(column = 2, row = cell[0].row).value
            item_repeat_day = [item_start_day]
        else:
            item_end_time = ws.cell(column = 2, row = cell[1].row).value
            item_repeat_day = list(range(item_start_day, item_start_day + cell[1].column - cell[0].column + 1))

        #print(cell, cell[0].value, item_start_time, item_end_time, item_start_day, item_repeat_day)
        event_info[cell] = {
            'val': cell[0].value, 
            'start_time_in_day': item_start_time, 
            'end_time_in_day': item_end_time, 
            #'start_day': item_start_day, 
            #'start_last_night': item_start_last_night, 
            'repeat_day': item_repeat_day
        }
    return start_monday, end_monday, event_info
if __name__ == "__main__":
    main()
